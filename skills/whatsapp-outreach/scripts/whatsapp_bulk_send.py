"""
WhatsApp 批量发送脚本
- 从 Excel 读取有号码的客户
- 根据行业和公司信息生成个性化消息
- 通过 wacli CLI 发送
- 安全控制：每条间隔 30-60 秒，确认后发送
- 内置 wacli LOCK 清理 + 进程清理 + 重试逻辑
"""
import openpyxl
import subprocess
import time
import random
import json
import re
import os
import sys

# ── 消息模板（按行业）───────────────────────────────────────

TEMPLATES = {
    '物流仓储': """Hi{name},

I'm Wike from Honglong Industrial — we manufacture conveyor belt processing equipment (slitting, splicing, skiving machines) in China.

I understand {company} works with conveyor systems in Saudi Arabia. Our machines help reduce belt maintenance costs by up to 40%.

Would you be open to reviewing our product catalog?

- Wike, Honglong Industrial""",

    '水泥建材': """Hi{name},

I'm Wike from Honglong Industrial. We specialize in conveyor belt splicing presses and processing equipment for heavy-duty applications.

I noticed {company} operates in the cement industry in Saudi Arabia. Our machines are built for demanding environments like cement plants.

Shall I share our latest catalog?

- Wike, Honglong Industrial""",

    '食品加工': """Hi{name},

I'm Wike from Honglong Industrial. We make conveyor belt processing equipment — splicing presses, slitting machines, and skiving tools.

I understand {company} handles food production with conveyor systems. Our equipment is used by food processors in 30+ countries.

Would you like to see our product range?

- Wike, Honglong Industrial""",

    '木工家具': """Hi{name},

I'm Wike from Honglong Industrial. We manufacture conveyor belt processing equipment including splicing presses and slitting machines.

I noticed {company} works in the wood industry in Saudi Arabia. Our machines help keep production lines running smoothly.

Shall I send you our catalog?

- Wike, Honglong Industrial""",

    'default': """Hi{name},

I'm Wike from Honglong Industrial — we manufacture conveyor belt processing equipment (slitting, splicing, skiving machines) in China.

I noticed {company} works with industrial systems in Saudi Arabia. We're looking for reliable partners in the region.

Would you be open to reviewing our product catalog?

- Wike, Honglong Industrial""",
}


# ── wacli 锁清理 ───────────────────────────────────────────

def clean_wacli_lock():
    """清理残留的 wacli LOCK 文件（sync 被 kill 后常会残留）"""
    lock_path = os.path.join(os.path.expanduser('~'), '.wacli', 'LOCK')
    try:
        if os.path.exists(lock_path):
            with open(lock_path) as f:
                content = f.read()
            pid_match = re.search(r'pid=(\d+)', content)
            if pid_match:
                pid = int(pid_match.group(1))
                try:
                    import psutil
                    if not psutil.pid_exists(pid):
                        os.remove(lock_path)
                        return True, "stale lock removed"
                except ImportError:
                    os.remove(lock_path)
                    return True, "lock removed (psutil unavailable)"
            return False, "lock file has no pid"
        return False, "no lock file"
    except Exception as e:
        return False, f"lock check failed: {e}"


def stop_wacli_processes():
    """停止所有 wacli 进程（避免 sync 和 send 冲突）"""
    try:
        subprocess.run(
            ['powershell', '-Command',
             'Stop-Process -Name wacli -Force -ErrorAction SilentlyContinue'],
            capture_output=True, timeout=10
        )
        time.sleep(2)
    except Exception:
        pass


# ── 发送逻辑 ───────────────────────────────────────────────

def generate_message(company_en, contact_name, industry):
    """生成个性化 WhatsApp 消息"""
    name = f" {contact_name}" if contact_name else ""
    template = TEMPLATES.get(industry, TEMPLATES['default'])
    msg = template.format(name=name, company=company_en)
    if len(msg) > 500:
        msg = msg[:497] + "..."
    return msg


def send_whatsapp(phone, message, max_retries=2):
    """通过 wacli 发送 WhatsApp 消息（自动处理锁问题）"""
    phone_clean = phone.replace('+', '').replace(' ', '').replace('-', '')
    jid = f"{phone_clean}@s.whatsapp.net"

    for attempt in range(max_retries):
        stop_wacli_processes()
        clean_wacli_lock()

        try:
            result = subprocess.run(
                ['wacli', 'send', 'text', '--to', jid, '--message', message],
                capture_output=True, text=True, timeout=30, encoding='utf-8'
            )
            if result.returncode == 0:
                return True, result.stdout.strip()

            err_msg = (result.stderr.strip() or result.stdout.strip())

            # 检测锁冲突 → 重试
            if 'lock' in err_msg.lower() or 'LOCK' in err_msg or result.returncode == 23:
                if attempt < max_retries - 1:
                    time.sleep(3)
                    continue
            return False, err_msg
        except subprocess.TimeoutExpired:
            return False, "Timeout (30s)"
        except Exception as e:
            return False, str(e)

    return False, "max retries exceeded"


def main():
    # 读取 Excel（路径通过 sys.argv 传入，或用默认路径）
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        excel_path = os.path.join(os.path.dirname(__file__), '..', '..',
                                  'WorkBuddy', 'latest', '沙特客户Pipeline.xlsx')
        excel_path = os.path.normpath(excel_path)

    if not os.path.exists(excel_path):
        # 尝试当前目录
        candidates = [
            '沙特客户Pipeline.xlsx',
            '沙特客户Pipeline_4行业.xlsx',
        ]
        for c in candidates:
            if os.path.exists(c):
                excel_path = c
                break
        else:
            print(f"❌ Excel 文件不存在: {excel_path}")
            sys.exit(1)

    results_path = excel_path.replace('.xlsx', '_whatsapp_results.json')

    wb = openpyxl.load_workbook(excel_path)
    sheet_names = [s for s in wb.sheetnames if 'Pipeline' in s or '沙特' in s]
    ws = wb[sheet_names[0]] if sheet_names else wb.active

    # 读取数据（列号需与实际 Excel 对应）
    targets = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        phone   = row[15].value   if len(row) > 15 else None
        company = row[4].value   if len(row) > 4 else None
        contact = row[12].value   if len(row) > 12 else None
        industry= row[1].value   if len(row) > 1 else 'default'
        icp     = row[16].value   if len(row) > 16 else None
        status  = row[18].value   if len(row) > 18 else None

        if not phone or not company:
            continue

        phone_str = str(phone).strip()
        if not phone_str.startswith('+966') and not phone_str.startswith('966') and not phone_str.startswith('05'):
            continue
        if not phone_str.startswith('+'):
            if phone_str.startswith('966'):
                phone_str = '+' + phone_str
            elif phone_str.startswith('05'):
                phone_str = '+966' + phone_str[1:]

        msg = generate_message(company, contact, industry)
        targets.append({
            'company': company,
            'contact': contact,
            'phone': phone_str,
            'industry': industry,
            'icp': icp,
            'status': status,
            'message': msg,
        })

    if not targets:
        print("❌ 没有找到可发送的目标号码")
        return

    print("=" * 70)
    print(f"📋 WhatsApp 批量发送预览 — 共 {len(targets)} 条")
    print("=" * 70)

    for i, t in enumerate(targets, 1):
        print(f"\n--- [{i}/{len(targets)}] {t['company']} ---")
        print(f"  📱 号码: {t['phone']}")
        print(f"  🏭 行业: {t['industry']} | ICP: {t['icp']} | 状态: {t['status']}")
        print(f"  📝 消息预览:")
        for line in t['message'].split('\n'):
            print(f"     {line}")

    print("\n" + "=" * 70)
    auto_confirm = '--yes' in sys.argv or '-y' in sys.argv
    if auto_confirm:
        confirm = 'yes'
        print("⚠️  自动确认发送 (--yes)")
    else:
        confirm = input("⚠️  确认发送以上消息？(yes/no): ").strip().lower()

    if confirm != 'yes':
        print("❌ 已取消发送")
        return

    results = []
    for i, t in enumerate(targets, 1):
        print(f"\n[{i}/{len(targets)}] 发送给 {t['company']} ({t['phone']})...")

        success, response = send_whatsapp(t['phone'], t['message'])

        result = {
            'company': t['company'],
            'phone': t['phone'],
            'success': success,
            'response': response,
            'timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
        }
        results.append(result)

        if success:
            print(f"  ✅ 发送成功: {response}")
        else:
            print(f"  ❌ 发送失败: {response}")

        # 更新 Excel 状态
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if len(row) > 4 and row[4].value == t['company']:
                try:
                    if len(row) > 18:
                        row[18].value = '已WhatsApp触达'
                except Exception:
                    pass
                break

        if i < len(targets):
            delay = random.randint(30, 60)
            print(f"  ⏳ 等待 {delay} 秒...")
            time.sleep(delay)

    wb.save(excel_path)
    with open(results_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    success_count = sum(1 for r in results if r['success'])
    print(f"\n{'=' * 70}")
    print(f"📊 发送完成：成功 {success_count}/{len(targets)}")
    print(f"💾 结果已保存: {results_path}")


if __name__ == '__main__':
    main()
