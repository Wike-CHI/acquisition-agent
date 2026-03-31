import openpyxl
import os
import glob

def read_excel_file(file_path, product_name):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        print(f'\n{"=" * 60}')
        print(f'{product_name}')
        print(f'{"=" * 60}')
        print(f'文件: {os.path.basename(file_path)}')
        print(f'工作表: {ws.title}')
        print(f'行数: {ws.max_row}, 列数: {ws.max_column}\n')
        
        # 读取前40行数据
        for row_idx in range(1, min(41, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(16, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                if cell.value is not None:
                    row_data.append(str(cell.value))
            if row_data:
                print(f'Row {row_idx}: {" | ".join(row_data)}')
    except Exception as e:
        print(f'错误: {e}')

# 读取分层机Word文档（如果有参数）
print('\n' + '=' * 60)
print('📐 分层机参数')
print('=' * 60)
print('\n分层机没有Excel参数表，参数在Word文档中')
print('需要手动从Word文档提取参数')

# 读取裁切机参数
print('\n' + '=' * 60)
print('✂️  裁切机参数')
print('=' * 60)

cutter_files = [
    ('temp-cutter-挡板裁切机参数.xlsx', '挡板裁切机'),
]

for file_name, product_name in cutter_files:
    if os.path.exists(file_name):
        read_excel_file(file_name, product_name)

# 读取导条机参数
print('\n' + '=' * 60)
print('🔩 导条机参数')
print('=' * 60)

guidestrip_files = [
    ('temp-guidestrip-挡板焊接机参数.xlsx', '挡板焊接机'),
    ('temp-guidestrip-C类（三代）导条机.xlsx', '三代导条机'),
]

for file_name, product_name in guidestrip_files:
    if os.path.exists(file_name):
        read_excel_file(file_name, product_name)
