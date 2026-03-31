import openpyxl
import json

# 读取风冷机价格表（本地副本）
wb = openpyxl.load_workbook('temp-windcool.xlsx')
ws = wb.active

print('========================================')
print('三代风冷机完整数据')
print('========================================\n')

# 读取所有数据
all_data = []
for row_idx in range(1, ws.max_row + 1):
    row_data = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, col_idx)
        if cell.value is not None:
            row_data.append(str(cell.value))
    if row_data:
        all_data.append(row_data)
        print(f'Row {row_idx}: {" | ".join(row_data)}')

print(f'\n总行数: {len(all_data)}')

# 查找价格信息
print('\n========================================')
print('查找价格信息')
print('========================================')
for idx, row in enumerate(all_data):
    if any('价格' in str(cell) or 'price' in str(cell).lower() or '元' in str(cell) or '$' in str(cell) for cell in row):
        print(f'Row {idx + 1}: {" | ".join(row)}')
