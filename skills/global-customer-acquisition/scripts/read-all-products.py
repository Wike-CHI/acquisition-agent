import openpyxl
import glob
import os

def read_excel_file(file_path, product_name):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        print(f'\n{"=" * 60}')
        print(f'{product_name}')
        print(f'{"=" * 60}')
        print(f'文件: {os.path.basename(file_path)}')
        print(f'工作表: {ws.title}')
        print(f'行数: {ws.max_row}, 列数: {ws.max_column}\n')
        
        # 读取前40行数据
        for row_idx in range(1, min(41, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(16, ws.max_column + 1)):
                cell = ws.cell(row_idx, col_idx)
                if cell.value is not None:
                    row_data.append(str(cell.value))
            if row_data:
                print(f'Row {row_idx}: {" | ".join(row_data)}')
    except Exception as e:
        print(f'错误: {e}')

# 读取水冷机参数
print('\n' + '=' * 60)
print('💧 水冷式接头机技术参数')
print('=' * 60)

watercool_files = [
    ('temp-watercool-7-二代水冷1100-2100X200.xlsx', '二代水冷 1100-2100X200'),
    ('temp-watercool-日式水冷参数表(1).xlsx', '日式水冷'),
    ('temp-watercool-7-水冷不锈钢1600X150以下.xlsx', '水冷不锈钢 1600X150'),
    ('temp-watercool-7-水冷铝合金2100-2600X150.xlsx', '水冷铝合金 2100-2600X150'),
]

for file_name, product_name in watercool_files:
    if os.path.exists(file_name):
        read_excel_file(file_name, product_name)

# 读取打齿机参数
print('\n' + '=' * 60)
print('🔧 打齿机技术参数')
print('=' * 60)

fingerpuncher_files = [
    ('temp-fingerpuncher-全自动打齿机参数.xlsx', '全自动打齿机'),
    ('temp-fingerpuncher-重型移动打齿机参数.xlsx', '重型移动打齿机'),
]

for file_name, product_name in fingerpuncher_files:
    if os.path.exists(file_name):
        read_excel_file(file_name, product_name)
