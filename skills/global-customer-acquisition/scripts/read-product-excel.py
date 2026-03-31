import openpyxl
import json

# 读取风冷机价格表（本地副本）
wb = openpyxl.load_workbook('temp-windcool.xlsx')

# 获取所有工作表名
print('工作表:', wb.sheetnames)
print('')

# 读取第一个工作表
ws = wb.active
print(f'工作表名: {ws.title}')
print(f'行数: {ws.max_row}')
print(f'列数: {ws.max_column}')
print('')

# 读取前30行数据
print('前30行数据:')
for row_idx in range(1, min(31, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(16, ws.max_column + 1)):
        cell = ws.cell(row_idx, col_idx)
        if cell.value is not None:
            row_data.append(str(cell.value))
    if row_data:
        print(f'Row {row_idx}: {" | ".join(row_data)}')
